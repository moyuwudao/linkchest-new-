import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

md_path = r'c:\Users\Mayn\CodeBuddy\20260407184558\translation-catalog.md'
output_path = r'c:\Users\Mayn\CodeBuddy\20260407184558\translation-catalog.xlsx'

with open(md_path, 'r', encoding='utf-8') as f:
    content = f.read()

lines = content.split('\n')

rows = []
current_source = None
current_namespace = None

for line in lines:
    stripped = line.strip()
    # Detect source sections
    if stripped.startswith('## 二、Web 端 UI 翻译'):
        current_source = 'web'
        current_namespace = None
        continue
    elif stripped.startswith('## 三、Mobile 端 UI 翻译'):
        current_source = 'mobile'
        current_namespace = None
        continue
    elif stripped.startswith('## 四、API 错误码消息'):
        current_source = 'api-error'
        current_namespace = None
        continue
    elif stripped.startswith('## 五、硬编码 Web 文本'):
        current_source = 'hardcoded-web'
        current_namespace = None
        continue
    elif stripped.startswith('## 六、硬编码 Mobile 文本'):
        current_source = 'hardcoded-mobile'
        current_namespace = None
        continue
    elif stripped.startswith('## 七、邮件模板文本'):
        current_source = 'email'
        current_namespace = None
        continue
    elif stripped.startswith('## '):
        current_source = None
        current_namespace = None
        continue

    if stripped.startswith('### '):
        current_namespace = stripped[4:].strip()
        continue

    if current_source and stripped.startswith('|') and 'Key' not in stripped and '---' not in stripped:
        parts = [p.strip() for p in stripped.split('|')]
        parts = [p for p in parts if p]
        if len(parts) >= 4:
            key = parts[0]
            zh = parts[1]
            en = parts[2]
            rows.append({
                'source': current_source,
                'namespace': current_namespace or '',
                'key': key,
                'zh': zh,
                'en': en,
            })
        elif len(parts) == 3 and current_source == 'api-error':
            # API error code tables have different format sometimes
            pass

# Parse API error codes separately since they use a slightly different format
api_section = re.search(r'## 四、API 错误码消息.*?(?=## 五、)', content, re.DOTALL)
if api_section:
    api_text = api_section.group(0)
    for line in api_text.split('\n'):
        stripped = line.strip()
        if stripped.startswith('|') and '错误码' not in stripped and '---' not in stripped:
            parts = [p.strip() for p in stripped.split('|')]
            parts = [p for p in parts if p]
            if len(parts) >= 3:
                key = parts[0]
                zh = parts[1]
                en = parts[2] if len(parts) > 2 else ''
                rows.append({
                    'source': 'api-error',
                    'namespace': 'errorCode',
                    'key': key,
                    'zh': zh,
                    'en': en,
                })

wb = Workbook()
ws = wb.active
ws.title = 'Translations'

# Headers
headers = ['来源(Source)', '分类(Namespace)', '键名(Key)', '中文(ZH)', '英文(EN)', '法语(FR)', '德语(DE)', '日语(JA)', '韩语(KO)', '备注(Context)']
ws.append(headers)

# Header style
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# Alternate row fills
fill_even = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

for i, r in enumerate(rows, start=2):
    ws.append([
        r['source'],
        r['namespace'],
        r['key'],
        r['zh'],
        r['en'],
        '', '', '', '', ''
    ])
    row_fill = fill_even if i % 2 == 0 else fill_odd
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=i, column=col)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# Column widths
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 45
ws.column_dimensions['F'].width = 40
ws.column_dimensions['G'].width = 40
ws.column_dimensions['H'].width = 40
ws.column_dimensions['I'].width = 40
ws.column_dimensions['J'].width = 30

# Freeze header
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = ws.dimensions

# Set row height for header
ws.row_dimensions[1].height = 30

wb.save(output_path)
print(f'Generated {output_path} with {len(rows)} rows')
